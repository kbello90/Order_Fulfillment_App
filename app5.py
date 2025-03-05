import streamlit as st
import pandas as pd
import pdfplumber
import io
import openpyxl
from openpyxl.utils import get_column_letter
from fpdf import FPDF

def extract_invoice_details(pdf_file):
    with pdfplumber.open(pdf_file) as pdf:
        text = "\n".join([page.extract_text() for page in pdf.pages if page.extract_text()])
    
    invoice_data = {
        "Sales Order No": "", "Date": "", "Bill To": "", "Ship To": "",
        "Items": [], "Importer of Record": "", "Special Instructions": "", "Incoterms": "", "Mode": "", "Freight Forwarder": "", "Total Weight": ""
    }
    
    return invoice_data

def generate_invoice_excel(invoice_data):
    template_path = "invoice_template2.xlsx"  # Updated pre-defined invoice template
    output = io.BytesIO()
    
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # Insert header details
    ws["I6"].value = invoice_data["Sales Order No"]
    ws["I7"].value = invoice_data["Date"]
    ws["B10"].value = invoice_data["Bill To"]
    ws["B14"].value = invoice_data["Ship To"]
    ws["B17"].value = invoice_data["Importer of Record"]
    ws["B25"].value = invoice_data["Special Instructions"]
    ws["B35"].value = invoice_data["Incoterms"]
    ws["B38"].value = invoice_data["Total Weight"]
    ws["B42"].value = invoice_data["Mode"]
    ws["B44"].value = invoice_data["Freight Forwarder"]
    
    # Insert items dynamically
    start_row = 30
    for index, item in enumerate(invoice_data["Items"]):
        row = start_row + index  # Place each item in the next available row
        ws[f"B{row}"].value = item["Description"]
        ws[f"G{row}"].value = item["Quantity"]
        ws[f"H{row}"].value = item["Unit Price"]
        ws[f"I{row}"].value = item["Total Price"]
    
    wb.save(output)
    output.seek(0)
    return output

def generate_invoice_pdf(invoice_data):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", style='B', size=14)
    pdf.cell(200, 10, "Commercial Invoice", ln=True, align='C')
    pdf.ln(10)
    pdf.set_font("Arial", size=12)
    
    for key, value in invoice_data.items():
        if key != "Items":
            pdf.cell(200, 10, f"{key}: {value}", ln=True)
    
    pdf.ln(10)
    pdf.cell(200, 10, "Items:", ln=True)
    for item in invoice_data["Items"]:
        pdf.cell(200, 10, f"{item['Description']} - Qty: {item['Quantity']} - Unit Price: {item['Unit Price']} - Total: {item['Total Price']}", ln=True)
    
    output = io.BytesIO()
    pdf.output(output, 'S')
    output.seek(0)
    return output.read()

# Streamlit App Interface
st.title("Commercial Invoice Generator")

option = st.radio("Choose Input Method:", ["Upload PDF", "Manual Entry"])

if option == "Upload PDF":
    uploaded_file = st.file_uploader("Upload Sales Order (PDF)", type=["pdf"])
    if uploaded_file is not None:
        invoice_data = extract_invoice_details(uploaded_file)
        st.write("Extracted Invoice Details:")
        st.json(invoice_data)
        excel_file = generate_invoice_excel(invoice_data)
        pdf_file = generate_invoice_pdf(invoice_data)
        
        st.download_button("Download Invoice (Excel)", data=excel_file.getvalue(), file_name="invoice.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.download_button("Download Invoice (PDF)", data=pdf_file, file_name="invoice.pdf", mime="application/pdf")

elif option == "Manual Entry":
    with st.form("invoice_form"):
        sales_order = st.text_input("Sales Order No")
        date = st.date_input("Invoice Date")
        bill_to = st.text_area("Bill To")
        ship_to = st.text_area("Ship To")
        importer = st.text_input("Importer of Record")
        special_instructions = st.text_area("Special Instructions")
        incoterms = st.text_input("Incoterms")
        mode = st.text_input("Mode")
        freight_forwarder = st.text_input("Freight Forwarder")
        total_weight = st.text_input("Total Weight (Kgs-Lbs)")
        
        items = []
        num_items = st.number_input("Number of Items", min_value=1, step=1)
        for i in range(num_items):
            description = st.text_area(f"Description {i+1}")
            quantity = st.number_input(f"Quantity {i+1}", min_value=1)
            unit_price = st.number_input(f"Unit Price {i+1}", min_value=0.0, format="%.2f")
            total_price = quantity * unit_price
            items.append({"Description": description, "Quantity": quantity, "Unit Price": unit_price, "Total Price": total_price})
        
        submit = st.form_submit_button("Generate Invoice")
        
    if submit:
        invoice_data = {
            "Sales Order No": sales_order,
            "Date": str(date),
            "Bill To": bill_to,
            "Ship To": ship_to,
            "Importer of Record": importer,
            "Special Instructions": special_instructions,
            "Incoterms": incoterms,
            "Mode": mode,
            "Freight Forwarder": freight_forwarder,
            "Total Weight": total_weight,
            "Items": items
        }
        excel_file = generate_invoice_excel(invoice_data)
        pdf_file = generate_invoice_pdf(invoice_data)
        
        st.download_button("Download Invoice (Excel)", data=excel_file.getvalue(), file_name="invoice.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.download_button("Download Invoice (PDF)", data=pdf_file, file_name="invoice.pdf", mime="application/pdf")

